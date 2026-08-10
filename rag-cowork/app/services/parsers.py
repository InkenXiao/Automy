"""文件解析器 · 按扩展名路由到对应解析通道

- pdf        : PyMuPDF 文本层 (电子版) → mineru 网关 (扫描件), 移植 pro-cowork doc_parse_service
- docx/pptx  : zip + XML 文本抽取 (无额外依赖)
- xlsx       : openpyxl 逐 sheet 抽取
- txt/md/csv : 直接读取
- 音频       : pydub 切片 + ASR 转写 (移植 pro-cowork asr_service)
- 图片       : VLM 内容描述

统一返回 {"text": str, "pages": int, "engine": str, "resources": [{resource_type, content_desc}]}
"""
import asyncio
import logging
import math
import os
import re
import tempfile
import zipfile
from pathlib import Path
from xml.etree import ElementTree as ET

import httpx

from app.config import settings
from app.services import llm_service

logger = logging.getLogger(__name__)

MIN_TEXT_CHARS = 50       # PDF 文本层低于该字符数视为扫描件
MINERU_TIMEOUT_S = 600    # mineru vLLM 冷启动 + 推理余量
MAX_CHUNK_RETRIES = 2

AUDIO_EXTS = {"mp3", "wav", "m4a", "aac", "flac", "ogg", "wma", "amr"}
IMAGE_EXTS = {"jpg", "jpeg", "png", "gif", "bmp", "webp", "tif", "tiff"}
TEXT_EXTS = {"txt", "md", "markdown", "csv", "json", "log", "html", "htm"}


# ---------------- PDF ----------------

def _pdf_pymupdf(path: Path) -> tuple[str, int]:
    import fitz

    parts: list[str] = []
    with fitz.open(str(path)) as doc:
        pages = doc.page_count
        for i, page in enumerate(doc, start=1):
            text = (page.get_text("text") or "").strip()
            if text:
                parts.append(f"----- 第 {i} 页 -----\n{text}")
    return "\n\n".join(parts), pages


async def _pdf_mineru(path: Path) -> str:
    if not settings.MINERU_API_URL:
        raise RuntimeError("mineru 算力网关未配置 (MINERU_API_URL)")
    url = settings.MINERU_API_URL.rstrip("/") + "/api/v1/parse/pdf"
    async with httpx.AsyncClient(timeout=MINERU_TIMEOUT_S) as client:
        with open(path, "rb") as f:
            resp = await client.post(url, files={"file": (path.name, f, "application/pdf")})
    if resp.status_code != 200:
        raise RuntimeError(f"mineru 网关返回 {resp.status_code}: {resp.text[:200]}")
    md = (resp.json().get("content") or "").strip()
    if not md:
        raise RuntimeError("mineru 网关返回空内容")
    return md


async def _parse_pdf(path: Path) -> dict:
    loop = asyncio.get_running_loop()
    try:
        text, pages = await loop.run_in_executor(None, _pdf_pymupdf, path)
    except ImportError as e:
        raise RuntimeError("缺少 PyMuPDF 依赖: pip install PyMuPDF") from e
    except Exception as e:  # noqa: BLE001
        raise RuntimeError(f"PDF 打开失败: {e}") from e

    if len(text) >= MIN_TEXT_CHARS:
        return {"text": text, "pages": pages, "engine": "pymupdf", "resources": []}

    if settings.MINERU_API_URL:
        try:
            md = await _pdf_mineru(path)
            if md:
                return {"text": md, "pages": pages, "engine": "mineru", "resources": []}
        except Exception as e:  # noqa: BLE001
            logger.warning("mineru 网关解析失败: %s", e)

    raise RuntimeError(
        "该 PDF 无文本层 (扫描件), 且 mineru 算力网关不可用, 请配置 MINERU_API_URL 后重试"
    )


# ---------------- Office (zip+XML / openpyxl) ----------------

def _xml_texts(data: bytes) -> list[str]:
    """从 XML 字节中抽取全部文本节点"""
    try:
        root = ET.fromstring(data)
    except ET.ParseError:
        return []
    return [t.strip() for t in root.itertext() if t and t.strip()]


def _parse_docx(path: Path) -> dict:
    parts: list[str] = []
    with zipfile.ZipFile(path) as z:
        names = [n for n in z.namelist() if re.fullmatch(r"word/(document|header\d*|footer\d*)\.xml", n)]
        for name in sorted(names):
            parts.extend(_xml_texts(z.read(name)))
    text = "\n".join(parts)
    if not text:
        raise RuntimeError("DOCX 未抽取到文本内容")
    return {"text": text, "pages": 0, "engine": "docx-xml", "resources": []}


def _parse_pptx(path: Path) -> dict:
    parts: list[str] = []
    with zipfile.ZipFile(path) as z:
        slides = sorted(
            (n for n in z.namelist() if re.fullmatch(r"ppt/slides/slide\d+\.xml", n)),
            key=lambda n: int(re.search(r"\d+", n).group()),
        )
        for i, name in enumerate(slides, start=1):
            texts = _xml_texts(z.read(name))
            if texts:
                parts.append(f"----- 第 {i} 页幻灯片 -----\n" + "\n".join(texts))
    text = "\n\n".join(parts)
    if not text:
        raise RuntimeError("PPTX 未抽取到文本内容")
    return {"text": text, "pages": len(slides), "engine": "pptx-xml", "resources": []}


def _parse_xlsx(path: Path) -> dict:
    from openpyxl import load_workbook

    wb = load_workbook(str(path), read_only=True, data_only=True)
    parts: list[str] = []
    for ws in wb.worksheets:
        rows: list[str] = []
        for row in ws.iter_rows(values_only=True):
            cells = [str(c) for c in row if c is not None and str(c).strip()]
            if cells:
                rows.append(" | ".join(cells))
        if rows:
            parts.append(f"----- 工作表: {ws.title} -----\n" + "\n".join(rows))
    wb.close()
    text = "\n\n".join(parts)
    if not text:
        raise RuntimeError("XLSX 未抽取到文本内容")
    return {"text": text, "pages": 0, "engine": "openpyxl", "resources": []}


# ---------------- 音频 (ASR) ----------------

def _fmt_ts(seconds: float) -> str:
    total = int(seconds)
    m, s = divmod(total, 60)
    h, m = divmod(m, 60)
    return f"[{h:02d}:{m:02d}:{s:02d}]" if h > 0 else f"[{m:02d}:{s:02d}]"


async def _asr_post_chunk(client: httpx.AsyncClient, wav_path: str, chunk_no: int) -> dict:
    last_err: Exception | None = None
    for attempt in range(MAX_CHUNK_RETRIES):
        try:
            with open(wav_path, "rb") as f:
                resp = await client.post(
                    settings.ASR_API_URL,
                    headers={"Authorization": f"Bearer {settings.ASR_API_KEY}"},
                    files={"file": f},
                    data={"model": settings.ASR_MODEL, "response_format": "verbose_json"},
                    timeout=httpx.Timeout(connect=10.0, read=300.0, write=60.0, pool=10.0),
                )
            if resp.status_code != 200:
                raise RuntimeError(f"HTTP {resp.status_code}: {resp.text[:200]}")
            return resp.json()
        except Exception as e:  # noqa: BLE001
            last_err = e
            if attempt + 1 < MAX_CHUNK_RETRIES:
                await asyncio.sleep(1)
    raise RuntimeError(f"音频片段 {chunk_no} 转录失败: {last_err}")


async def _parse_audio(path: Path) -> dict:
    """音频切片转写为带时间戳文本 (移植 pro-cowork asr_service)"""
    if not settings.ASR_API_URL:
        raise RuntimeError("ASR 未配置: 请设置 ASR_API_URL / ASR_API_KEY")
    try:
        from pydub import AudioSegment
    except ImportError as e:
        raise RuntimeError("缺少 pydub 依赖, 且容器内需有 ffmpeg") from e

    loop = asyncio.get_running_loop()
    try:
        audio = await loop.run_in_executor(None, AudioSegment.from_file, str(path))
    except Exception as e:  # noqa: BLE001
        raise RuntimeError(f"音频加载失败 (需 ffmpeg 支持该格式): {e}") from e

    total_ms = len(audio)
    chunk_ms = max(int(settings.ASR_CHUNK_MS), 10000)
    total_chunks = max(1, math.ceil(total_ms / chunk_ms))

    segments: list[dict] = []
    async with httpx.AsyncClient() as client:
        for i in range(total_chunks):
            chunk = audio[i * chunk_ms: min((i + 1) * chunk_ms, total_ms)]
            tmp = tempfile.NamedTemporaryFile(prefix=f"asr_chunk_{i}_", suffix=".wav", delete=False)
            tmp_path = tmp.name
            tmp.close()
            try:
                await loop.run_in_executor(None, lambda c=chunk, p=tmp_path: c.export(p, format="wav"))
                data = await _asr_post_chunk(client, tmp_path, i + 1)
            finally:
                if os.path.exists(tmp_path):
                    os.remove(tmp_path)
            chunk_segments = data.get("segments") or []
            if chunk_segments:
                for seg in chunk_segments:
                    text = (seg.get("text") or "").strip()
                    if text:
                        segments.append({"start": (seg.get("start") or 0) + i * chunk_ms / 1000, "text": text})
            else:
                text = (data.get("text") or "").strip()
                if text:
                    segments.append({"start": i * chunk_ms / 1000, "text": text})

    lines = [f"{_fmt_ts(s['start'])} {s['text']}" for s in segments]
    text = "\n".join(lines)
    if not text:
        raise RuntimeError("ASR 转写结果为空")
    return {
        "text": text,
        "pages": 0,
        "engine": "asr",
        "resources": [{"resource_type": "audio", "content_desc": f"语音转写 (时长 {round(total_ms / 1000, 1)}s): {text[:500]}"}],
    }


# ---------------- 图片 (VLM) ----------------

async def _parse_image(path: Path) -> dict:
    data = path.read_bytes()
    desc = await llm_service.vision_describe(data)
    if not desc:
        raise RuntimeError("VLM 图片识别失败或未配置 VISION 通道")
    return {
        "text": f"[图片] {path.name}\n{desc}",
        "pages": 1,
        "engine": "vlm",
        "resources": [{"resource_type": "image", "content_desc": desc}],
    }


# ---------------- 统一入口 ----------------

async def parse_file(file_path: str | Path) -> dict:
    """按扩展名解析文件, 返回 {"text", "pages", "engine", "resources"}"""
    path = Path(file_path)
    if not path.exists():
        raise RuntimeError(f"文件不存在: {path.name}")
    ext = path.suffix.lower().lstrip(".")

    if ext == "pdf":
        return await _parse_pdf(path)
    if ext == "docx":
        return await asyncio.to_thread(_parse_docx, path)
    if ext == "pptx":
        return await asyncio.to_thread(_parse_pptx, path)
    if ext in ("xlsx", "xlsm"):
        return await asyncio.to_thread(_parse_xlsx, path)
    if ext in TEXT_EXTS:
        text = path.read_text(encoding="utf-8", errors="ignore").strip()
        if not text:
            raise RuntimeError("文件内容为空")
        return {"text": text, "pages": 0, "engine": "text", "resources": []}
    if ext in AUDIO_EXTS:
        return await _parse_audio(path)
    if ext in IMAGE_EXTS:
        return await _parse_image(path)

    raise RuntimeError(f"暂不支持的文件类型: .{ext} (支持 pdf/office/音频/图片/文本)")
