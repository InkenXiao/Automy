"""项目周报 Excel 导出服务 (需求: 项目经理导出全部项目周报)

严格复刻参照文件 demo/20260807_数字智能项目工作周报_肖立军.xlsx:
- 单 Sheet (名 pm{经理}), 每项目 13 行区块:
  标题(合并A:G 微软雅黑16加粗) → 项目名称/经理/周期 → 项目状态/总工时(公式)/总体进度
  → 空行 → 深蓝表头(FF1F4E78 白字12pt) → 5 个成员槽位(10pt, 斑马纹 FFF2F2F2)
  → 本周问题汇总(FFD9E1F2) → 项目总体概况(FFD9E1F2) → 空行
- 工作任务列按天聚合: 「周一：xxx（参与人员：yyy）；」换行拼接
- 列宽 A8.7/B14.7/D64.5/E19.7/F16.5/G56.5; 周期显示 周一~周五
"""
import math
from datetime import date, timedelta
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from sqlalchemy import func, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.personal_report import PersonalReport
from app.models.progress_task import ProgressTask
from app.models.project import Project
from app.models.project_member import ProjectMember
from app.models.weekly_report import WeeklyReport

# ---------- 样式常量 (取自参照文件) ----------
FONT = "微软雅黑"
HEADER_FILL = PatternFill("solid", fgColor="FF1F4E78")
ZEBRA_FILL = PatternFill("solid", fgColor="FFF2F2F2")
SUMMARY_FILL = PatternFill("solid", fgColor="FFD9E1F2")
THIN = Side(style="thin")
DOUBLE = Side(style="double")
BORDER_THIN = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
BORDER_SUMMARY = Border(left=THIN, right=THIN, top=THIN, bottom=DOUBLE)
COL_WIDTHS = {"A": 8.7, "B": 14.7, "D": 64.5, "E": 19.7, "F": 16.5, "G": 56.5}
HEADERS = ["序号", "项目成员", "项目角色", "本周具体工作任务", "交付物 / 产出", "本周工时 (h)", "下周计划安排"]
DAY_NAMES = {1: "周一", 2: "周二", 3: "周三", 4: "周四", 5: "周五", 6: "周六", 7: "周日"}
MEMBER_SLOTS = 5
BLOCK_ROWS = 13  # 每项目区块行数 (含首尾空行)


def _font(size: int, bold: bool = False, color: str = "FF000000") -> Font:
    return Font(name=FONT, size=size, bold=bold, color=color)


def _align(h: str = "center", wrap: bool = True) -> Alignment:
    return Alignment(horizontal=h, vertical="center", wrap_text=wrap)


def _est_height(pairs: list[tuple[str, float]], min_h: float) -> float:
    """按 (文本, 每行容纳汉字数) 估算行高; 换行符分段累计"""
    lines = 1
    for text, chars_per_line in pairs:
        if not text:
            continue
        n = 0
        for seg in str(text).split("\n"):
            n += max(1, math.ceil(len(seg) / max(chars_per_line, 1)))
        lines = max(lines, n)
    return max(min_h, lines * 13.5 + 4)


def _set(ws, coord: str, value, font: Font, align: Alignment, fill=None, border=None):
    """写单元格并套用样式"""
    c = ws[coord]
    c.value = value
    c.font = font
    c.alignment = align
    if fill:
        c.fill = fill
    if border:
        c.border = border
    return c


def _style_range(ws, cell_range: str, fill=None, border=None):
    """对合并区域全部单元格补 fill/border (openpyxl 合并区仅左上角生效显示, 边框需逐格)"""
    for row in ws[cell_range]:
        for c in row:
            if fill:
                c.fill = fill
            if border:
                c.border = border


async def _member_rows(db: AsyncSession, project: Project, week_start: date) -> list[dict]:
    """装配成员槽位数据: 该项目该周个人周报 → 按天聚合工作任务/交付物/工时/下周计划"""
    reports = (
        (
            await db.execute(
                select(PersonalReport)
                .where(
                    PersonalReport.is_delete.is_(False),
                    PersonalReport.project_id == project.id,
                    PersonalReport.week_start == week_start,
                )
                .order_by(PersonalReport.id)
            )
        )
        .scalars()
        .all()
    )
    if not reports:
        return []
    report_ids = [r.id for r in reports]

    from app.models.personal_report import PersonalReportPlanItem, PersonalReportWorkItem

    work_items = (
        (
            await db.execute(
                select(PersonalReportWorkItem)
                .where(
                    PersonalReportWorkItem.is_delete.is_(False),
                    PersonalReportWorkItem.report_id.in_(report_ids),
                )
                .order_by(
                    PersonalReportWorkItem.report_id,
                    PersonalReportWorkItem.day_of_week,
                    PersonalReportWorkItem.sort_order,
                    PersonalReportWorkItem.id,
                )
            )
        )
        .scalars()
        .all()
    )
    plan_items = (
        (
            await db.execute(
                select(PersonalReportPlanItem)
                .where(
                    PersonalReportPlanItem.is_delete.is_(False),
                    PersonalReportPlanItem.report_id.in_(report_ids),
                )
                .order_by(PersonalReportPlanItem.report_id, PersonalReportPlanItem.sort_order, PersonalReportPlanItem.id)
            )
        )
        .scalars()
        .all()
    )

    # 成员角色与排序 (项目经理排首行)
    members = (
        (
            await db.execute(
                select(ProjectMember)
                .where(ProjectMember.is_delete.is_(False), ProjectMember.project_id == project.id)
                .order_by(ProjectMember.sort_order, ProjectMember.id)
            )
        )
        .scalars()
        .all()
    )
    role_map = {m.name: (m.role or "") for m in members}
    order_map = {m.name: i for i, m in enumerate(members)}

    rows: list[dict] = []
    for report in reports:
        # 工作任务: 仅取属于本项目(或未指定项目)的行, 按天聚合
        by_day: dict[int, list[str]] = {}
        deliverables: list[str] = []
        total_hours = 0.0
        for w in work_items:
            if w.report_id != report.id:
                continue
            if w.project_id not in (None, project.id):
                continue
            total_hours += w.hours or 0
            content = (w.content or "").strip()
            if content:
                if w.participants:
                    content += f"（参与人员：{w.participants}）"
                by_day.setdefault(w.day_of_week or 1, []).append(content)
            dv = (w.deliverable or "").strip()
            if dv and dv not in deliverables:
                deliverables.append(dv)
        day_lines = []
        for d in sorted(by_day):
            day_lines.append(f"{DAY_NAMES.get(d, '周一')}：{'；'.join(by_day[d])}；")
        # 下周计划: 仅本项目行, 多条编号
        plans = [
            (p.content or "").strip()
            for p in plan_items
            if p.report_id == report.id and p.project_id in (None, project.id) and (p.content or "").strip()
        ]
        if len(plans) > 1:
            plan_text = "\n".join(f"{i}、{c}" for i, c in enumerate(plans, 1))
        else:
            plan_text = plans[0] if plans else ""

        is_pm = (project.manager or "").strip() == report.member_name
        rows.append({
            "name": report.member_name,
            "role": role_map.get(report.member_name) or ("项目经理" if is_pm else ""),
            "work": "\n".join(day_lines),
            "deliverable": "、".join(deliverables),
            "hours": round(total_hours, 2),
            "plan": plan_text,
            "is_pm": is_pm,
            "order": order_map.get(report.member_name, 999),
        })

    # 项目经理置顶, 其余按项目成员顺序
    rows.sort(key=lambda r: (not r["is_pm"], r["order"]))
    return rows[:MEMBER_SLOTS]


async def _project_block_data(db: AsyncSession, project: Project, week_start: date) -> dict:
    """装配单项目区块数据"""
    # 该周项目周报 (取最新一份)
    report = (
        (
            await db.execute(
                select(WeeklyReport)
                .where(
                    WeeklyReport.is_delete.is_(False),
                    WeeklyReport.project_id == project.id,
                    WeeklyReport.week_start == week_start,
                )
                .order_by(WeeklyReport.id.desc())
                .limit(1)
            )
        )
        .scalars()
        .first()
    )

    # 本周问题汇总 = 周报风险逐行
    problems = ""
    if report:
        from app.models.weekly_report import WeeklyRisk

        risks = (
            (
                await db.execute(
                    select(WeeklyRisk)
                    .where(WeeklyRisk.is_delete.is_(False), WeeklyRisk.report_id == report.id)
                    .order_by(WeeklyRisk.sort_order, WeeklyRisk.id)
                )
            )
            .scalars()
            .all()
        )
        problems = "\n".join(r.title for r in risks if (r.title or "").strip())

    # 总体进度 = 进度计划完成率 (done / 总数)
    total = (
        await db.execute(
            select(ProgressTask.id).where(
                ProgressTask.is_delete.is_(False), ProgressTask.project_id == project.id
            )
        )
    ).all()
    done = (
        await db.execute(
            select(ProgressTask.id).where(
                ProgressTask.is_delete.is_(False),
                ProgressTask.project_id == project.id,
                ProgressTask.status == "done",
            )
        )
    ).all()
    pct = round(len(done) / len(total) * 100) if total else 0

    week_end = week_start + timedelta(days=4)  # 周期显示 周一~周五 (同参照文件)
    return {
        "title": f"【项目工作周报】 {project.name}",
        "project_title": project.title or project.name,
        "manager": project.manager or "",
        "period": f"        周报周期：{week_start:%Y/%m/%d} - {week_end:%Y/%m/%d}",
        "status": project.status or "",
        "progress": f"        总体进度：{pct}％",
        "members": await _member_rows(db, project, week_start),
        "problems": problems,
        "overview": (report.week_digest or report.overview_summary or "") if report else "",
    }


def _write_block(ws, r0: int, data: dict) -> None:
    """写单项目 13 行区块 (r0 为 1-based 起始行)"""
    # 1. 标题 (合并 A:G, 16pt 加粗居中)
    ws.merge_cells(f"A{r0}:G{r0}")
    _set(ws, f"A{r0}", data["title"], _font(16, bold=True), _align())
    ws.row_dimensions[r0].height = 22.85

    # 2. 项目名称 / 项目经理 / 周报周期
    ws.merge_cells(f"A{r0+1}:B{r0+1}")
    ws.merge_cells(f"C{r0+1}:D{r0+1}")
    _set(ws, f"A{r0+1}", "项目名称：", _font(12, bold=True), _align())
    _set(ws, f"C{r0+1}", data["project_title"], _font(12), _align("left"))
    _set(ws, f"E{r0+1}", "项目经理：", _font(12, bold=True), _align())
    _set(ws, f"F{r0+1}", data["manager"], _font(12), _align())
    _set(ws, f"G{r0+1}", data["period"], _font(12, bold=True), _align("left"))
    ws.row_dimensions[r0 + 1].height = 16.9

    # 3. 项目状态 / 本周总工时(公式) / 总体进度
    first_slot = r0 + 5
    last_slot = r0 + 9
    formula = "=" + "+".join(f"F{r}" for r in range(first_slot, last_slot + 1))
    ws.merge_cells(f"A{r0+2}:B{r0+2}")
    ws.merge_cells(f"C{r0+2}:D{r0+2}")
    _set(ws, f"A{r0+2}", "项目状态：", _font(12, bold=True), _align())
    _set(ws, f"C{r0+2}", data["status"], _font(12), _align("left"))
    _set(ws, f"E{r0+2}", "本周总工时：", _font(12, bold=True), _align())
    _set(ws, f"F{r0+2}", formula, _font(12), _align())
    _set(ws, f"G{r0+2}", data["progress"], _font(12, bold=True), _align("left"))
    ws.row_dimensions[r0 + 2].height = 16.9

    # 4. (r0+3) 空行; 5. 深蓝表头
    hr = r0 + 4
    for i, h in enumerate(HEADERS):
        col = chr(ord("A") + i)
        _set(
            ws, f"{col}{hr}", h,
            _font(12, bold=True, color="FFFFFFFF"), _align(),
            fill=HEADER_FILL, border=BORDER_THIN,
        )
    ws.row_dimensions[hr].height = 16.9

    # 6. 成员槽位 (固定 5 行; 斑马纹: 第 2/4 行整行灰, 第 1 行 D 列灰 — 与参照一致)
    members = data["members"]
    for i in range(MEMBER_SLOTS):
        r = r0 + 5 + i
        m = members[i] if i < len(members) else None
        zebra = ZEBRA_FILL if i in (1, 3) else None
        values = [
            i + 1,
            m["name"] if m else "",
            m["role"] if m else "",
            m["work"] if m else "",
            m["deliverable"] if m else "",
            (m["hours"] if m and m["hours"] else None),
            m["plan"] if m else "",
        ]
        for j, col in enumerate("ABCDEFG"):
            align = _align("left") if col in ("D", "E", "G") else _align()
            fill = zebra or (ZEBRA_FILL if (i == 0 and col == "D") else None)
            _set(ws, f"{col}{r}", values[j], _font(10), align, fill=fill, border=BORDER_THIN)
        if m:
            ws.row_dimensions[r].height = _est_height(
                [(m["work"], 30), (m["deliverable"], 9), (m["plan"], 26)], 13.85
            )
        else:
            ws.row_dimensions[r].height = 13.85

    # 7. 本周问题汇总 / 8. 项目总体概况 (A:B 合并标签 + C:G 合并内容, 浅蓝底, 底部双线)
    for offset, (label, content, min_h) in enumerate([
        ("本周问题汇总", data["problems"], 28.0),
        ("项目总体概况", data["overview"], 40.0),
    ]):
        r = r0 + 10 + offset
        ws.merge_cells(f"A{r}:B{r}")
        ws.merge_cells(f"C{r}:G{r}")
        _set(ws, f"A{r}", label, _font(12, bold=True), _align())
        _set(ws, f"C{r}", content, _font(10, bold=True), _align("left"))
        _style_range(ws, f"A{r}:G{r}", fill=SUMMARY_FILL, border=BORDER_SUMMARY)
        ws.row_dimensions[r].height = _est_height([(content, 78)], min_h)
    # 9. (r0+12) 空行


async def build_pm_weekly_excel(
    db: AsyncSession, manager_name: str, week_start: date
) -> tuple[BytesIO, str] | None:
    """导出该项目经理全部项目周报 Excel

    返回 (buffer, filename); 该用户不是任何项目的经理时返回 None
    """
    projects = (
        (
            await db.execute(
                select(Project)
                .where(
                    Project.is_delete.is_(False),
                    func.trim(Project.manager) == manager_name,
                )
                .order_by(Project.sort_order, Project.id)
            )
        )
        .scalars()
        .all()
    )
    if not projects:
        return None

    wb = Workbook()
    ws = wb.active
    ws.title = f"pm{manager_name}"
    for col, width in COL_WIDTHS.items():
        ws.column_dimensions[col].width = width

    r0 = 1
    for project in projects:
        data = await _project_block_data(db, project, week_start)
        _write_block(ws, r0, data)
        r0 += BLOCK_ROWS

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    filename = f"{week_start:%Y%m%d}_数字智能项目工作周报_{manager_name}.xlsx"
    return buffer, filename
