"""任务/对话附件 → 提示词片段 共享构建器

供 routers/task_runs.py (长任务) 与 routers/agents.py (数字分身对话/调试) 共用:
- 音频: 指引 Agent 经 run_skill 调用「会议纪要生成」技能
- 图片: 指引 Agent 经 run_skill 调用「图像识别」技能 (视觉多模态模型)
- PDF : 指引 Agent 经 run_skill 调用「文档解析」技能 (PyMuPDF/mineru/paddleocr)
- Office (xlsx/docx/pptx): 本地确定性解析为文本后内联 (openpyxl / zip+XML, 不走技能)
- 其他文本类附件 (含 html): 直接内联文件内容 (截断至 MAX_FILE_CONTENT_CHARS)
"""
import re
import zipfile
from pathlib import Path
from typing import Optional
from xml.etree import ElementTree as ET

# 任务附件存储目录 (与 routers/task_runs.py 一致): pro-cowork/data/task_files/<project_id>/<filename>
TASK_FILES_ROOT = Path(__file__).resolve().parent.parent.parent / "data" / "task_files"
MAX_FILE_CONTENT_CHARS = 20000  # 注入提示词的文件内容上限

AUDIO_EXTS = {".mp3", ".wav", ".m4a", ".ogg", ".flac", ".aac", ".wma", ".opus"}
IMAGE_EXTS = {".png", ".jpg", ".jpeg", ".gif", ".webp", ".bmp", ".svg"}
PDF_EXTS = {".pdf"}
OFFICE_EXTS = {".xlsx", ".xlsm", ".docx", ".pptx"}


def file_kind(filename: str) -> str:
    """附件类别: audio / image / pdf / office / text"""
    ext = Path(filename).suffix.lower()
    if ext in AUDIO_EXTS:
        return "audio"
    if ext in IMAGE_EXTS:
        return "image"
    if ext in PDF_EXTS:
        return "pdf"
    if ext in OFFICE_EXTS:
        return "office"
    return "text"


def safe_filename(name: str) -> str:
    """去除路径分隔符, 防目录穿越"""
    return Path(name).name


# ---------------- Office 确定性解析 (zip+XML / openpyxl, 无重模型依赖) ----------------

def _xml_texts(data: bytes) -> list[str]:
    """从 XML 字节中抽取全部文本节点"""
    try:
        root = ET.fromstring(data)
    except ET.ParseError:
        return []
    return [t.strip() for t in root.itertext() if t and t.strip()]


def _parse_docx_text(path: Path) -> str:
    parts: list[str] = []
    with zipfile.ZipFile(path) as z:
        names = [n for n in z.namelist() if re.fullmatch(r"word/(document|header\d*|footer\d*)\.xml", n)]
        for name in sorted(names):
            parts.extend(_xml_texts(z.read(name)))
    return "\n".join(parts)


def _parse_pptx_text(path: Path) -> str:
    parts: list[str] = []
    with zipfile.ZipFile(path) as z:
        slides = sorted(
            (n for n in z.namelist() if re.fullmatch(r"ppt/slides/slide\d+\.xml", n)),
            key=lambda n: int(re.search(r"\d+", n).group()),
        )
        for i, name in enumerate(slides, start=1):
            texts = _xml_texts(z.read(name))
            if texts:
                parts.append(f"----- 第 {i} 页幻灯片 -----\n" + "\n".join(texts))
    return "\n\n".join(parts)


def _parse_xlsx_text(path: Path) -> str:
    from openpyxl import load_workbook

    wb = load_workbook(str(path), read_only=True, data_only=True)
    parts: list[str] = []
    for ws in wb.worksheets:
        rows: list[str] = []
        for row in ws.iter_rows(values_only=True):
            cells = [str(c) for c in row if c is not None and str(c).strip()]
            if cells:
                rows.append(" | ".join(cells))
        if rows:
            parts.append(f"----- 工作表: {ws.title} -----\n" + "\n".join(rows))
    wb.close()
    return "\n\n".join(parts)


def parse_office_text(path: Path) -> str:
    """Office 附件 → 纯文本; 失败抛出 RuntimeError"""
    ext = path.suffix.lower()
    if ext == ".docx":
        text = _parse_docx_text(path)
    elif ext == ".pptx":
        text = _parse_pptx_text(path)
    elif ext in (".xlsx", ".xlsm"):
        text = _parse_xlsx_text(path)
    else:
        raise RuntimeError(f"不支持的 Office 类型: {ext}")
    if not text.strip():
        raise RuntimeError(f"{path.name} 未抽取到文本内容")
    return text


def read_task_file_text(project_id: Optional[int], filename: str) -> str:
    """读取附件文本内容 (office 先确定性解析; 音频/图片/PDF 不内联, 由对应技能处理)"""
    path = TASK_FILES_ROOT / str(project_id or 0) / safe_filename(filename)
    if not path.exists():
        return ""
    kind = file_kind(filename)
    try:
        if kind == "office":
            return parse_office_text(path)[:MAX_FILE_CONTENT_CHARS]
        if kind != "text":
            return ""
        return path.read_text(encoding="utf-8", errors="ignore")[:MAX_FILE_CONTENT_CHARS]
    except Exception:
        return ""


def build_file_prompt_parts(
    project_id: Optional[int], file_names: list[str], skill_guidance: bool = True
) -> list[str]:
    """将附件列表转换为提示词片段列表

    skill_guidance=True : 音频/图片/PDF 给技能调用指引 (模型自主调 run_skill 的场景)
    skill_guidance=False: 音频/图片/PDF 跳过指引 (调用方已做确定性预处理, 见 attachment_service)
    """
    parts: list[str] = []
    for fname in file_names or []:
        fname = safe_filename(fname)
        kind = file_kind(fname)
        pid = project_id or 0
        if kind in ("audio", "image", "pdf") and not skill_guidance:
            continue  # 由 attachment_service 预处理注入解析结果
        if kind == "audio":
            parts.append(
                f"【录音文件 {fname}】音频附件, 请通过 run_skill 调用「会议纪要生成」技能处理 "
                f"(input_data: {{\"file_name\": \"{fname}\", \"project_id\": {pid}}}), "
                f"并将转写文字与生成的会议纪要完整展示"
            )
        elif kind == "image":
            parts.append(
                f"【图片文件 {fname}】图片附件, 请通过 run_skill 调用「图像识别」技能处理 "
                f"(input_data: {{\"file_name\": \"{fname}\", \"project_id\": {pid}}}), "
                f"并将识别出的图片内容完整展示"
            )
        elif kind == "pdf":
            parts.append(
                f"【PDF文件 {fname}】文档附件, 请通过 run_skill 调用「文档解析」技能处理 "
                f"(input_data: {{\"file_name\": \"{fname}\", \"project_id\": {pid}}}), "
                f"并将解析出的文档内容完整展示"
            )
        else:
            content = read_task_file_text(project_id, fname)
            if content:
                parts.append(f"【附件 {fname}】\n{content}")
            elif kind == "office":
                parts.append(
                    f"【附件 {fname}】Office 文件本地解析失败 (文件可能损坏或格式不受支持), "
                    f"请提示用户转换为 PDF 或文本后重新上传"
                )
    return parts
